# -*- coding: utf-8 -*-
"""
Created on Fri Nov  5 21:32:38 2021

@author: Yz Wu
"""

import urllib.request
import re
from bs4 import BeautifulSoup


import openpyxl
from openpyxl.styles import PatternFill


import tkinter as tk 
from tkinter import filedialog
import tkinter.messagebox


import time
import datetime


import random




#建立窗口window
window = tk.Tk()
 
#给窗口的可视化起名字
window.title('1688价格查询')
 
#设定窗口的大小(长 x 宽)
window.geometry('500x370')  
#设定程序图标
window.iconbitmap('1688.ico')


## 设定标签
#使用说明
img = tk.PhotoImage(file = 'warning.png')
label_img = tk.Label(window,image = img)
label_img.place(x=85,y=10)

#文件位置
fileLocation = tk.Label(window, text="", bg='white', fg='black', font=('楷体', 12), width=50, height=2)
fileLocation.place(x=50, y=250)

#点击按钮导入信息表
l_priceList = tk.Label(window, text="点击按钮导入信息表", bg='green', fg='white', font=('楷体', 14), width=20, height=2)
# 说明： bg为背景，fg为字体颜色，font为字体，width为长，height为高，这里的长和高是字符的长和高，比如height=2,就是标签有2个字符高
l_priceList.place(x=50, y=190)




##定义Function，既点击按钮时执行的程序
#导入信息表按钮
def hit_me_p():

    l_priceList.configure(text='信息表已导入',bg='red')
    #全局变量，获取装箱单路径
    global priceList
    priceList = filedialog.askopenfilename()
    fileLocation.configure(text = priceList)

    
##开始按钮
def generate():
    
    #获取开始时间
    time_start = time.time()
  
    
    #信息表workbook
    wb = openpyxl.load_workbook(priceList)
    #定位信息表workbook中的工作表worksheet
    sheet = wb.worksheets[0]
    #获取最大行
    maxRow = sheet.max_row
    
    
    # 从第三行到最大行循环
    for i in range(3,maxRow+1):
        print(i)
        
        #随机等待时间
        waitTime = random.randint(1,15)

        #获取网站地址
        urlpage = sheet['E'+str(i)].value
        
        if urlpage == None:
            continue
                
        # 读取页面信息
        page = urllib.request.urlopen(urlpage)
        
        # 使用BeautifulSoup优化信息
        soup = BeautifulSoup(page.read(), 'lxml')
        
        #定位价格
        finder = re.findall(r'priceRange.*', str(soup))
        # print(finder)
        
        #如果链接失效
        if finder == []:
            # 输出错误
            sheet['R'+str(i)] = '错误'
            sheet['S'+str(i)] = '错误'
            sheet['T'+str(i)] = '错误'
            
            #标红
            sheet['R'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
            sheet['S'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
            sheet['T'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    
            #保存
            wb.save('test.xlsx')
        else:
            #获取价格
            priceRange=re.findall(r'\d+\.?\d*',finder[0])
            
            if len(priceRange)==6:
                #填写价格
                sheet['R'+str(i)] = priceRange[1]
                sheet['S'+str(i)] = priceRange[3]
                sheet['T'+str(i)] = priceRange[5]
                            
                     
                #如果价格有变动，标红    
                if sheet['N'+str(i)].value - float(priceRange[1]) != 0.0:
                    sheet['R'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
                
                if sheet['O'+str(i)].value - float(priceRange[3]) != 0.0:
                    sheet['S'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
                
                if sheet['P'+str(i)].value - float(priceRange[5]) != 0.0:
                    sheet['T'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    
                #保存
                wb.save(priceList)
                
                #等待
                time.sleep(waitTime)
            elif len(priceRange)==4:
                sheet['R'+str(i)] = priceRange[1]
                sheet['S'+str(i)] = priceRange[3]
                                 
                     
                #如果价格有变动，标红    
                if sheet['N'+str(i)].value - float(priceRange[1]) != 0.0:
                    sheet['R'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
                
                if sheet['O'+str(i)].value - float(priceRange[3]) != 0.0:
                    sheet['S'+str(i)].fill=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
                
                #保存
                wb.save(priceList)
                
                #等待
                time.sleep(waitTime)
                


        #关闭工作表
        wb.close()


    #获取结束时间    
    time_end = time.time()
    #计算运行时间
    runningTime = str(datetime.timedelta(seconds = round(time_end-time_start)))  
          
    
    #弹出消息框，完成并通报运行时间
    tkinter.messagebox.showinfo(title='完成', message='恭喜您！价格更新完毕！用时'+runningTime)


## 放置Button
#导入信息表按钮
b_priceList = tk.Button(window, text='1.导入信息表', font=('楷体', 14), width=13, height=1, command=hit_me_p)
b_priceList.place(x=317, y=200)


#开始更新按钮
b_template = tk.Button(window, text='2.开始更新', font=('楷体', 14), width=12, height=1, command=generate)
b_template.place(x=175, y=310)

 

#主窗口循环显示
window.mainloop()


















